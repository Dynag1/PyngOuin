import fichier.var as var
import fichier.design as design
import threading
from tkinter import filedialog
from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook
name = ""

def chSave():
    Tk().withdraw()
    filename = filedialog.asksaveasfilename(initialdir="/", title="Select file", filetypes=(
    ("Excel files", "*.xlsx"), ("all files", "*.*")))
    return filename

def chOpen():
    filename = filedialog.askopenfilename(initialdir="/", title="Select File", filetypes=(
    ("Excel files", "*.xlsx"), ("all files", "*.*")))
    return filename

def saveExcel():

    name = chSave()
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "IP"
    sheet["b1"] = "Nom"
    sheet["C1"] = "Mac"
    sheet["D1"] = "Port"
    sheet["E1"] = "Latence"
    sheet["F1"] = "Comm"
    x = 2
    for parent in var.app_instance.tab_ip.get_children():
        result = var.app_instance.tab_ip.item(parent)["values"]
        try:
            resultComm = result[6]
        except:
            resultComm = ""
        sheet["A" + str(x)] = str(result[0])
        sheet["b" + str(x)] = str(result[1])
        sheet["C" + str(x)] = str(result[2])
        sheet["D" + str(x)] = str(result[3])
        sheet["E" + str(x)] = str(result[4])
        sheet["F" + str(x)] = str(resultComm)
        x = x + 1
    workbook.save(filename=name + ".xlsx")
    threading.Thread(target=design.alert, args=("Votre fichier excel à bien été crée",)).start()

def openExcel():
    name = chOpen()
    design.tab_erase()

    workbook = load_workbook(filename=name)
    workbook.sheetnames
    sheet = workbook.active
    sheet.title
    x = 2
    for row in sheet.rows:
        for value in sheet.iter_rows(min_row=x, max_row=x, values_only=True):
            ipexist = False
            for parent in var.app_instance.tab_ip.get_children():
                result = var.app_instance.tab_ip.item(parent)["values"]
                ip1 = value[0]
                ip = result[0]

                if ip1 == ip:
                    threading.Thread(target=design.alert, args=("L'adresse " + ip + " existe déja",)).start()
                    ipexist = True
                    pass
                else:
                    ipexist = False
            if ipexist == False:
                mac = ""
                port = ""
                comm = ''
                if value[1] != None:
                    nom = value[1]
                else:
                    nom = value[0]
                if value[2] != None:
                    mac = value[2]
                if value[3] != None:
                    port = value[3]
                if value[5] != None:
                    comm = value[5]
                var.app_instance.tab_ip.insert(parent='', index=x, tag=value[0], iid=value[0],
                                                values=(value[0], nom, mac, port, "", "", comm))
        x += 1